import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from colorama import Fore, init
from datetime import datetime
import platform
import os
import re
init(autoreset=True)
def create_form_descriptor() -> None:
  try:
    wb = Workbook()
  except Exception as e:
    print(Fore.RED + f'Error initiating Workbook: {e}')
    return
  ws1 = wb.active
  ws1.title = 'Tech-stacks'
  ws1.append(['Tecnologia', 'Pergunta de Familiaridade', 'Pergunta de Interesse', 'Valores possíveis', 'Opções de Uso', 'Finalidade da Tecnologia'])
  levels = '0 a 5'
  q1 = 'Qual é o seu nível de familiaridade com '
  q2 = 'Qual é o seu nível de interesse em aprender/trabalhar com '
  techs = ['Planilhamento','Redação de Documentos', 
           'Softwares de Planejamento de Recursos de Negócios', 'Softwares de Gerenciamento de Relação com Clientes', 
           'Aplicativos de Gestão de Atividades e Planejamento', 'Aplicativos de Inteligência de Negócios',
           'Construtores de Formulários', 'Aplicativos de Armazenamento em Nuvem',
           'Grande Modelos de Linguagem', 'Inteligências Artificiais Generativas de Imagem', 
           'Inteligências Artificiais Generativas de Vídeo', 'Inteligências Artificiais Generativas de Áudio']
  tech_data = [
        [
            f"{techs[0]} (Microsoft Excel, Google Sheets, Libre Office Calc, etc.)",
            f"{q1}{techs[0]}?",
            f"{q2}{techs[0]}?",
            levels,
            "Fórmulas Avançadas | Tabelas Dinâmicas | Macros/VBA | Gráficos e Dashboards | Validação de Dados | Colaboração em Tempo Real",
            "Organizar e analisar dados de forma estruturada"
        ],
        [
            f"{techs[1]} (Microsoft Word, Google Docs, Libre Office Writer, etc.)",
            f"{q1}{techs[1]}?",
            f"{q2}{techs[1]}?",
            levels,
            "Modelos e Templates | Referências Cruzadas | Mala Direta | Estilos e Formatação Automática | Exportação para PDF",
            "Criar, editar e formatar documentos"
        ],
        [
            f"{techs[2]} (SAP, SAT, TOTVS, Salesforce, ERPGo, etc.)",
            f"{q1}{techs[2]}?",
            f"{q2}{techs[2]}?",
            levels,
            "Gestão de Estoques | Controle Financeiro | Relatórios de Faturamento | Monitoramento de Processos | Gestão de Produção",
            "Gerenciar processos empresariais em áreas como finanças e logística."
        ],
        [
            f"{techs[3]} (Monday.com, ClickUp, Slack, Jira, etc.)",
            f"{q1}{techs[3]}?",
            f"{q2}{techs[3]}?",
            levels,
            "Automação de Vendas | Gestão de Relacionamento | Relatórios de Análise | Integrações com Email | Segmentação de Clientes",
            "Acompanhar funil de vendas e relacionamento com clientes."
        ],
        [
            f"{techs[4]} (Notion, Trello, Microsoft Planner, Google Calendar, etc.)",
            f"{q1}{techs[4]}?",
            f"{q2}{techs[4]}?",
            levels,
            "Gestão de Tarefas | Criação de Workflows | Colaboração em Tempo Real | Monitoramento de Progresso",
            "Organizar tarefas, projetos e equipes de forma ágil."
        ],
        [
            f"{techs[5]} (PowerBI, Qlik Sense, Tableau, etc.)",
            f"{q1}{techs[5]}?",
            f"{q2}{techs[5]}?",
            levels,
            "Painéis Interativos | Visualização de Dados | Integração com múltiplas fontes",
            "Suportar análise de dados corporativos e tomada de decisão."
        ],
        [
            f"{techs[6]} (Google Forms, Jotform, Typeform, etc.)",
            f"{q1}{techs[6]}?",
            f"{q2}{techs[6]}?",
            levels,
            "Criação de Formulários | Coleta de Respostas | Integração com planilhas",
            "Facilitar pesquisa, feedback e coleta de dados."
        ],
        [
            f"{techs[7]} (Google Drive, Dropbox, Amazon S3, etc.)",
            f"{q1}{techs[7]}?",
            f"{q2}{techs[7]}?",
            levels,
            "Armazenamento e Compartilhamento de Arquivos | Backup e Sincronização",
            "Disponibilizar arquivos e dados na nuvem de forma segura."
        ],
        [
            f"{techs[8]} — LLMs (ChatGPT, GitHub Copilot, Gemini, LLaMa, etc.)",
            f"{q1}{techs[8]}?",
            f"{q2}{techs[8]}?",
            levels,
            "Geração de Relatórios | Criação de Scripts e Templates | Análise e Resumo de Textos | Tradução Automática | Geração de Código",
            "Automatizar e otimizar tarefas de redação ou análise de texto."
        ],
        [
            f"{techs[9]} (Dall-E, Midjourney, Stable Diffusion, etc.)",
            f"{q1}{techs[9]}?",
            f"{q2}{techs[9]}?",
            levels,
            "Criação de Imagens Publicitárias | Edição de Fotos | Prototipação artística | Finalidades Artísticas Diversas",
            "Produzir imagens estáticas com base em prompts textuais."
        ],
        [
            f"{techs[10]} (Sora, Runway, Fliki, etc.)",
            f"{q1}{techs[10]}?",
            f"{q2}{techs[10]}?",
            levels,
            "Geração de Vídeos Promocionais | Apresentações Audiovisuais | Alimentação de Conteúdos em Mídias Sociais | Finalidades Artísticas Diversas",
            "Produzir ou editar vídeos com base em prompts textuais ou referências."
        ],
        [
            f"{techs[11]} (ElevenLabs, PlayHT, ParrotAI, etc.)",
            f"{q1}{techs[11]}?",
            f"{q2}{techs[11]}?",
            levels,
            "Acessibilidade | Apresentações Audiovisuais | Finalidades Artísticas Diversas",
            "Gerar ou transformar áudio/voz de forma automatizada."
        ]
    ]
  try:
    for r in tech_data: ws1.append(r)
  except Exception as e:
    print(Fore.ORANGE + f'Error appending Technologies Data: {e}')
  ws2 = wb.create_sheet('Geral')
  ws2.append(['Nome do Dado', 'Tipo do Dado', 'Obrigatório', 'Filtros do Dado'])
  general_data = [
    [
      'Nome',
      'Texto',
      'Sim',
      'Não aceita números ou caracteres especiais'
    ],
    [
      'Sobrenome',
      'Texto',
      'Sim',
      'Não aceita números ou caracteres especiais'
    ],
    [
      'Idade',
      'Número',
      'Sim',
      'Máximo 127 | Somente dígitos numéricos',  
    ],
    [
      'Gênero',
      'Texto',
      'Sim',
      'Masculino | Feminino | Não-binário | Outros'
    ],
    [
      'E-mail Primário',
      'Texto',
      'Sim',
      'Mínimo 8 caracteres | Máximo 16383 caracteres | Deve incluir @ e .'
    ],
    [
      'E-mail Secundário',
      'Texto',
      'Não',
      'Mínimo 8 caracteres | Máximo 16383 caracteres | Deve incluir @ e .'
    ],
    [
      'Telefone',
      'Inteiro',
      'Sim',
      'Só aceita números, hífens e +'
    ],
    [
      'Telefone Secundário',
      'Inteiro',
      'Não',
      'Só aceita números, hífens e +'
    ],
    [
      "Cargo",
      "Texto",
      'Sim',
      "Executivo/Administrativo | Gerência/Direção | Supervisão | Financeiro | Compras | Suporte | Operações | Desenvolvimento"
    ],
    [
      "Tempo na Empresa (em meses)"
      "Número",
      'Sim',
      "Somente números"
    ]
  ]
  try:
    for rg in general_data: ws2.append(rg)
  except Exception as e:
    print(Fore.ORANGE + f'Error appending Technologies Data: {e}')
  header_fill = PatternFill(start_color='CCEFCB', end_color='CCEFCB', fill_type='solid')
  even_fill = PatternFill(start_color='EEEEEE', end_color='EEEEEE', fill_type='solid')
  odd_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
  def style_worksheet(sheet) -> None:
    try:
      header_row = sheet[1]
    except Exception as e:
      print(Fore.ORANGE + f'Error accessing sheet header. Aborting proccess.')
      return
    thin_border = Border(left=Side(style='thin', color='DDDDDD'), right=Side(style='thin', color='DDDDDD'))
    for c in header_row:
      c.fill = header_fill
      c.font = Font(bold=True)
      c.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
    max_col = sheet.max_column
    try:
      for ri in range(2, sheet.max_row + 1):
        fill = even_fill if (ri % 2 == 0) else odd_fill
        for ci in range(1, max_col + 1):
          curr_cell = sheet.cell(row=ri, column=ci)
          curr_cell.fill = fill
          max_width = 0
          max_length = 0
          if curr_cell.value:
            max_length = max(max_length, len(str(curr_cell.value)))
            max_width = max_length * 1.25
          if ci > 1 and ci < sheet.max_row: 
            curr_cell.border = thin_border
          col_desc = sheet.column_dimensions[get_column_letter(ci)]
          col_desc.wrap_text = False
          col_desc.width = max(max_width, col_desc.width)
    except IndexError:
      print(Fore.ORANGE + f'There was an error with index processing for the {sheet.title}')
    except Exception as e:
      print(Fore.RED + f'An undefined error stopped the execution of the loop for styling {sheet.title}: {e}')
  for ws in [ws1, ws2]: 
    style_worksheet(ws)
  try:
    wb.save(f'Prestech_form_techs_experiences_{re.sub(r'\s', '__', os.getlogin())}_{str(datetime.now()).split(' ')[0]}.xlsx')
    print(Fore.GREEN + 'The generation of the spreadsheet completed successfully.')
  except PermissionError:
    print(Fore.RED + f'Permission to save the file denied by {os.name} {platform.platform()}.')
  except Exception as e:
    print(Fore.RED + f'An undefined error stopped the execution: {e}')

if __name__ == '__main__':
  create_form_descriptor()
else: 
  print(Fore.RED + 'Directory not authorized for execution')